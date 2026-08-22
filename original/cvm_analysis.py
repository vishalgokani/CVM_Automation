#!/usr/bin/env python3
"""
CVM analyses -- one self-contained, shareable script.
====================================================
Runs the two analyses that back the CVM doming-ratio pipeline and writes everything to
`analysis_output/`. Deterministic: same inputs -> same numbers.

  1. METHOD BENCHMARK  -- quad-vs-mask IoU of the shipped Douglas-Peucker corner method
     against two off-the-shelf alternatives (quadrilateral-fitter, min-area rectangle),
     over every annotated vertebra. This is the evidence for choosing Douglas-Peucker.

  2. AGREEMENT VALIDATION -- Bland-Altman of the automated C2 doming ratio vs the expert
     (Mason) by-hand measurement: bias, 95% limits of agreement, Pearson r, Lin's CCC,
     ICC(2,1), MAE/RMSE, plus a Bland-Altman plot and a scatter-vs-identity plot.

It also writes an Excel copy of the measurements (with the ratio convention documented).

Ratio convention (denominator is always a POSTERIOR height):
    C2 doming ratio = C2 dome / C3 posterior      (C2's own height includes the dens)
    C3 doming ratio = C3 dome / C3 posterior
    C4 doming ratio = C4 dome / C4 posterior

Requirements: numpy, opencv-python, matplotlib, openpyxl.  Optional: quadrilateral-fitter
(the benchmark skips it, with a note, if it is not installed).  Needs the deployable
run_cvm.py (path configured below) and the annotated study folder to recompute quads.

Usage:   python3 cvm_analysis.py
"""
import os, sys, csv, glob, json, re, zipfile
import xml.etree.ElementTree as ET
from concurrent.futures import ProcessPoolExecutor, as_completed
import numpy as np

# ------------------------------------------------------------------ configuration
ROOT   = os.path.dirname(os.path.abspath(__file__))
DEPLOY = os.path.join(ROOT, "Deploy")                       # holds run_cvm.py
TEMP   = os.path.join(ROOT, "temp")                         # Batch*/output.json + images
CSV    = os.path.join(ROOT, "FINAL", "measurements.csv")    # measurements (DP, correct ratios)
MASON  = os.path.join(ROOT, "mason measurement updates.xlsx")
OUT    = os.path.join(ROOT, "analysis_output")
NS     = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

sys.path.insert(0, DEPLOY)
import run_cvm as rc                                        # the deployable pipeline
import cv2
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt


# =============================================================== shared helpers
def load_tasks():
    """One entry per annotated image: its annotations, category map, size, spine axis."""
    rc.index_images(TEMP, exclude=[os.path.join(TEMP, "cvm_output")])
    tasks = []
    for bj in glob.glob(os.path.join(TEMP, "Batch*", "output.json")):
        d = json.load(open(bj)); cats = {c["id"]: c["name"] for c in d.get("categories", [])}
        by = {}
        for a in d.get("annotations", []):
            by.setdefault(a["image_id"], []).append(a)
        for im in d.get("images", []):
            anns = by.get(im["id"], [])
            cents = [np.asarray(rc.polygons_of(a)[0], float).reshape(-1, 2).mean(0)
                     for a in anns if rc.polygons_of(a)]
            if not cents:
                continue
            tasks.append({"anns": anns, "cats": cats, "H": im["height"], "W": im["width"],
                          "axis": rc.spine_axis(cents).tolist()})
    return tasks


def iou(mask, quad):
    ys, xs = np.where(mask)
    if len(xs) == 0:
        return 0.0
    x0, y0 = max(0, xs.min() - 2), max(0, ys.min() - 2); x1, y1 = xs.max() + 2, ys.max() + 2
    m = mask[y0:y1, x0:x1].astype(bool)
    z = np.zeros_like(m, np.uint8)
    cv2.fillPoly(z, [np.round(np.asarray(quad) - [x0, y0]).astype(np.int32)], 1)
    return (m & z.astype(bool)).sum() / max(1, (m | z.astype(bool)).sum())


# =============================================================== 1. method benchmark
def _has_qfitter():
    try:
        import quadrilateral_fitter  # noqa: F401
        return True
    except Exception:
        return False


def _bench_worker(t):
    """Per-image: return [(category, dp_iou, qf_iou, minrect_iou), ...] for its vertebrae."""
    out = []
    have_qf = _has_qfitter()
    if have_qf:
        from quadrilateral_fitter import QuadrilateralFitter
    for a in t["anns"]:
        polys = rc.polygons_of(a)
        if not polys:
            continue
        try:
            cn = t["cats"].get(a["category_id"])
            P = np.vstack([np.asarray(p, float).reshape(-1, 2) for p in polys])
            mask = rc.poly_to_mask(polys, t["H"], t["W"])
            dp = iou(mask, rc.fit_quads(polys, np.array(t["axis"]), category=cn)["option_c"])
            mrq = iou(mask, cv2.boxPoints(cv2.minAreaRect(P.astype(np.float32))).astype(np.float64))
        except Exception:
            continue
        qfv = np.nan
        if have_qf:
            try:
                f = QuadrilateralFitter(polygon=P); f.fit()          # fit() BEFORE reading corners
                q = np.array(f.fitted_quadrilateral, dtype=float)
                if q.shape == (4, 2):
                    qfv = iou(mask, q)
            except Exception:
                pass
        out.append((cn, dp, qfv, mrq))
    return out


def benchmark_methods():
    have_qf = _has_qfitter()
    tasks = load_tasks()
    res = []
    with ProcessPoolExecutor(max_workers=max(1, (os.cpu_count() or 2) - 1)) as ex:
        for fut in as_completed([ex.submit(_bench_worker, t) for t in tasks]):
            res.extend(fut.result())
    cats = np.array([r[0] for r in res])
    cur = np.array([r[1] for r in res]); qf = np.array([r[2] for r in res]); mr = np.array([r[3] for r in res])
    lines = ["METHOD BENCHMARK -- quad-vs-mask IoU (higher = tighter fit)",
             "=" * 62,
             f"n = {len(cur)} vertebrae" + ("" if have_qf else "   [quadrilateral-fitter NOT installed -- skipped]"),
             f"  Douglas-Peucker (shipped) : mean {np.nanmean(cur):.3f}   median {np.nanmedian(cur):.3f}",
             f"  quadrilateral-fitter      : mean {np.nanmean(qf):.3f}   median {np.nanmedian(qf):.3f}",
             f"  min-area rectangle        : mean {np.nanmean(mr):.3f}   median {np.nanmedian(mr):.3f}",
             "  by vertebra (Douglas-Peucker):"]
    for v in ["C2", "C3", "C4"]:
        m = cats == v
        lines.append(f"    {v}: DP={np.nanmean(cur[m]):.3f}  qfit={np.nanmean(qf[m]):.3f}  minrect={np.nanmean(mr[m]):.3f}")
    txt = "\n".join(lines)
    open(os.path.join(OUT, "method_benchmark.txt"), "w").write(txt + "\n")
    print(txt + "\n")


# =============================================================== 2. Bland-Altman validation
def _mason_hand():
    z = zipfile.ZipFile(MASON)
    ss = [(t.text or "") for t in ET.fromstring(z.read("xl/sharedStrings.xml")).iter(NS + "t")]
    rows = {}
    for c in ET.fromstring(z.read("xl/worksheets/sheet1.xml")).iter(NS + "c"):
        v = c.find(NS + "v")
        if v is None:
            continue
        col = re.match(r"[A-Z]+", c.get("r")).group(); rn = int(re.search(r"\d+", c.get("r")).group())
        rows.setdefault(rn, {})[col] = ss[int(v.text)] if c.get("t") == "s" else v.text
    out = {}
    for r in rows.values():                          # B = vertebra, G = hand C2 ratio
        if r.get("B") == "C2" and r.get("G"):
            try:
                out[r["A"]] = float(r["G"])
            except ValueError:
                pass
    return out


def _automated_c2():
    out = {}
    for r in csv.DictReader(open(CSV)):
        if r["vertebra"] == "C2" and r["doming_ratio"]:
            out[r["scan"]] = float(r["doming_ratio"])
    return out


def _icc21(x, y):
    M = np.column_stack([x, y]).astype(float); n, k = M.shape; gm = M.mean()
    msr = k * ((M.mean(1) - gm) ** 2).sum() / (n - 1)
    msc = n * ((M.mean(0) - gm) ** 2).sum() / (k - 1)
    mse = ((M - M.mean(1, keepdims=True) - M.mean(0, keepdims=True) + gm) ** 2).sum() / ((n - 1) * (k - 1))
    return (msr - mse) / (msr + (k - 1) * mse + k * (msc - mse) / n)


def _ccc(x, y):
    x, y = np.asarray(x), np.asarray(y)
    return (2 * np.cov(x, y, ddof=0)[0, 1]) / (x.var() + y.var() + (x.mean() - y.mean()) ** 2)


def validate_bland_altman():
    hand, auto = _mason_hand(), _automated_c2()
    scans = sorted(set(hand) & set(auto))
    h = np.array([hand[s] for s in scans]); a = np.array([auto[s] for s in scans])
    diff, mean = a - h, (a + h) / 2
    bias, sd = diff.mean(), diff.std(ddof=1)
    lo, hi = bias - 1.96 * sd, bias + 1.96 * sd
    r = np.corrcoef(h, a)[0, 1]
    keep = np.abs(diff) <= 0.10
    lines = ["AGREEMENT VALIDATION -- automated C2 ratio vs Mason hand (Bland-Altman)",
             "=" * 62,
             f"n paired C2 scans = {len(scans)}",
             f"  bias (auto - hand)      : {bias:+.4f}",
             f"  95% limits of agreement : {lo:+.4f} to {hi:+.4f}",
             f"  Pearson r               : {r:.3f}",
             f"  Lin's CCC               : {_ccc(h, a):.3f}",
             f"  ICC(2,1)                : {_icc21(h, a):.3f}",
             f"  MAE / RMSE              : {np.abs(diff).mean():.4f} / {np.sqrt((diff**2).mean()):.4f}",
             f"  excl. |diff|>0.10 (n={keep.sum()}): bias {diff[keep].mean():+.4f}, MAE {np.abs(diff[keep]).mean():.4f}",
             "  NOTE: overall stats are outlier-driven (bad masks + cases the model corrects);",
             "        a blinded expert re-read with n>28 is needed for a publishable ICC."]
    txt = "\n".join(lines)
    open(os.path.join(OUT, "bland_altman.txt"), "w").write(txt + "\n")
    print(txt + "\n")

    with open(os.path.join(OUT, "bland_altman_data.csv"), "w", newline="") as f:
        w = csv.writer(f); w.writerow(["scan", "mason_hand", "automated", "diff", "mean"])
        for s, hi_, ai_ in zip(scans, h, a):
            w.writerow([s, round(hi_, 3), round(ai_, 3), round(ai_ - hi_, 3), round((hi_ + ai_) / 2, 3)])

    plt.figure(figsize=(7, 5))
    plt.scatter(mean, diff, s=40, color="#2b6cb0", zorder=3)
    plt.axhline(bias, color="#c53030", lw=1.5, label=f"bias {bias:+.3f}")
    plt.axhline(hi, color="#718096", ls="--", lw=1.2, label=f"+1.96 SD {hi:+.3f}")
    plt.axhline(lo, color="#718096", ls="--", lw=1.2, label=f"-1.96 SD {lo:+.3f}")
    plt.axhline(0, color="#000", lw=0.6, alpha=0.4)
    plt.xlabel("Mean of automated & hand C2 ratio"); plt.ylabel("Difference (automated - hand)")
    plt.title(f"Bland-Altman (n={len(scans)})"); plt.legend(fontsize=8); plt.tight_layout()
    plt.savefig(os.path.join(OUT, "bland_altman.png"), dpi=140); plt.close()

    plt.figure(figsize=(6, 6)); lim = [0, max(h.max(), a.max()) * 1.1]
    plt.plot(lim, lim, color="#718096", ls="--", lw=1, label="identity")
    plt.scatter(h, a, s=40, color="#2b6cb0", zorder=3)
    plt.xlim(lim); plt.ylim(lim); plt.xlabel("Mason hand C2 ratio"); plt.ylabel("Automated C2 ratio")
    plt.title(f"Automated vs hand (r={r:.3f}, CCC={_ccc(h,a):.3f})")
    plt.legend(fontsize=9); plt.gca().set_aspect("equal"); plt.tight_layout()
    plt.savefig(os.path.join(OUT, "scatter.png"), dpi=140); plt.close()


# =============================================================== excel export
def write_excel():
    import openpyxl
    from openpyxl.styles import Font
    rows = list(csv.reader(open(CSV)))
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "measurements"
    ws.append([("CVM doming measurements (Douglas-Peucker).  Denominator is a posterior height:  "
                "C2 = C2 dome / C3 posterior;  C3 = C3 dome / C3 posterior;  C4 = C4 dome / C4 posterior.")])
    ws["A1"].font = Font(italic=True)                          # row 1 = title
    ws.append(rows[0])                                         # row 2 = header
    for i in range(len(rows[0])):
        ws.cell(2, i + 1).font = Font(bold=True)
    for r in rows[1:]:                                         # row 3+ = data
        ws.append([float(x) if re.fullmatch(r"-?\d+(\.\d+)?", x or "") else x for x in r])
    for col, w in zip("ABCDEF", (46, 10, 20, 16, 13, 15)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    dest = os.path.join(os.path.dirname(CSV), "measurements.xlsx")
    wb.save(dest); print("wrote", dest)


def main():
    os.makedirs(OUT, exist_ok=True)
    print(f"Writing analyses to {OUT}\n")
    write_excel()
    benchmark_methods()
    validate_bland_altman()
    print("Done.")


if __name__ == "__main__":
    main()
